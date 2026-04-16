"""Document and file tools: file management, PDF, spreadsheets, charts."""

import os
import json
import csv
import re
import subprocess
import sys
import tempfile
import html as html_module

from langchain_core.tools import Tool
from langchain_community.agent_toolkits import FileManagementToolkit
from pypdf import PdfReader
from fpdf import FPDF
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from config import SANDBOX_DIR


def get_file_tools():
    """Return LangChain file management tools rooted in the sandbox."""
    toolkit = FileManagementToolkit(root_dir=SANDBOX_DIR)
    return toolkit.get_tools()


def read_pdf(file_path: str) -> str:
    """Read text content from a PDF file in the sandbox directory."""
    full_path = os.path.join(SANDBOX_DIR, file_path)
    if not os.path.isfile(full_path):
        return f"Error: file not found at {full_path}"
    reader = PdfReader(full_path)
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append(f"--- Page {i + 1} ---\n{text}")
    if not pages:
        return "The PDF contains no extractable text."
    return "\n\n".join(pages)


def _content_to_html(title: str, content: str) -> str:
    """Convert title + plain-text content into a styled HTML document."""
    escaped_title = html_module.escape(title)

    body_lines = []
    for line in content.split("\n"):
        if not line.strip():
            body_lines.append("<br>")
            continue
        safe = html_module.escape(line)
        safe = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", safe)
        if re.match(r"^\s*[-*]\s", safe):
            safe = re.sub(r"^\s*[-*]\s", "• ", safe)
        body_lines.append(f"<p style='margin:2px 0'>{safe}</p>")

    body_html = "\n".join(body_lines)

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><style>
  body {{ font-family: -apple-system, Arial, Helvetica, sans-serif; margin: 0; padding: 40px;
         font-size: 11pt; line-height: 1.6; color: #222; }}
  h1 {{ font-size: 18pt; margin: 0 0 16px 0; }}
</style></head><body>
{"<h1>" + escaped_title + "</h1>" if title else ""}
{body_html}
</body></html>"""


def _create_pdf_playwright(html_content: str, full_path: str) -> str | None:
    """Try to create a PDF using Playwright/Chromium. Returns error string or None on success."""
    tmp_html_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".html", delete=False, mode="w", encoding="utf-8") as tmp:
            tmp.write(html_content)
            tmp_html_path = tmp.name

        abs_pdf_path = os.path.abspath(full_path)
        script = (
            "from playwright.sync_api import sync_playwright\n"
            "with sync_playwright() as p:\n"
            "    browser = p.chromium.launch()\n"
            "    page = browser.new_page()\n"
            f"    page.goto('file://{tmp_html_path}')\n"
            f"    page.pdf(path=r'{abs_pdf_path}', format='A4',"
            "     margin={'top':'20mm','bottom':'20mm','left':'15mm','right':'15mm'})\n"
            "    browser.close()\n"
        )

        result = subprocess.run(
            [sys.executable, "-c", script],
            capture_output=True, text=True, timeout=120,
        )

        if result.returncode != 0:
            return result.stderr[-500:]
        return None
    except Exception as e:
        return str(e)
    finally:
        if tmp_html_path:
            try:
                os.unlink(tmp_html_path)
            except Exception:
                pass


def _sanitize_for_fpdf(text: str) -> str:
    """Replace Unicode characters that cause issues with fpdf built-in fonts."""
    replacements = {
        "\u2014": "--", "\u2013": "-", "\u2018": "'", "\u2019": "'",
        "\u201c": '"', "\u201d": '"', "\u2026": "...", "\u2022": "-",
        "\u00a0": " ", "\u200b": "",
    }
    for char, repl in replacements.items():
        text = text.replace(char, repl)
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _create_pdf_fpdf(title: str, content: str, full_path: str) -> str | None:
    """Create a PDF using fpdf (fallback). Returns error string or None on success."""
    try:
        title = _sanitize_for_fpdf(title)
        content = _sanitize_for_fpdf(content)

        pdf = FPDF()
        pdf.set_margins(15, 15, 15)
        pdf.add_page()

        if title:
            pdf.set_font("Helvetica", style="B", size=16)
            pdf.multi_cell(0, 10, title, align="L")
            pdf.ln(4)

        pdf.set_font("Helvetica", size=11)
        for paragraph in content.split("\n"):
            if paragraph.strip() == "":
                pdf.ln(4)
            else:
                pdf.multi_cell(0, 7, paragraph)

        pdf.output(full_path)
        return None
    except Exception as e:
        return str(e)


def create_pdf(input: str) -> str:
    """Create a proper PDF file in the sandbox directory from a title and text content.
    Input must be a JSON string with keys: 'filename', 'title' (optional), 'content'.
    Example: {"filename": "report.pdf", "title": "My Report", "content": "Text here..."}
    """
    try:
        data = json.loads(input)
    except json.JSONDecodeError as e:
        return f"Error: input must be valid JSON. {e}"

    filename = data.get("filename", "output.pdf")
    title = data.get("title", "")
    content = data.get("content", "")

    if not filename.endswith(".pdf"):
        filename += ".pdf"

    full_path = os.path.join(SANDBOX_DIR, filename)
    os.makedirs(SANDBOX_DIR, exist_ok=True)

    html_content = _content_to_html(title, content)
    pw_error = _create_pdf_playwright(html_content, full_path)
    if pw_error is None:
        return f"PDF successfully created at {SANDBOX_DIR}/{filename}"

    fpdf_error = _create_pdf_fpdf(title, content, full_path)
    if fpdf_error is None:
        return f"PDF successfully created at {SANDBOX_DIR}/{filename}"

    return f"Error creating PDF: {fpdf_error}"


def read_spreadsheet(file_path: str) -> str:
    """Read a CSV or Excel file from the sandbox directory and return a summary with the first rows.
    Pass the file path relative to the sandbox folder (e.g. 'data.csv' or 'report.xlsx').
    """
    full_path = os.path.join(SANDBOX_DIR, file_path)
    if not os.path.isfile(full_path):
        return f"Error: file not found at {full_path}"

    ext = os.path.splitext(file_path)[1].lower()
    rows = []
    headers = []

    try:
        if ext == ".csv":
            with open(full_path, newline="", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    if i == 0:
                        headers = row
                    rows.append(row)
        elif ext in (".xlsx", ".xls"):
            wb = openpyxl.load_workbook(full_path, read_only=True, data_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                str_row = [str(c) if c is not None else "" for c in row]
                if i == 0:
                    headers = str_row
                rows.append(str_row)
            wb.close()
        else:
            return f"Error: unsupported file type '{ext}'. Use .csv, .xlsx, or .xls."
    except Exception as e:
        return f"Error reading file: {e}"

    total_rows = len(rows)
    preview_rows = rows[:21]

    lines = [f"File: {file_path}  |  {total_rows} rows  |  {len(headers)} columns"]
    lines.append(f"Columns: {', '.join(headers)}")
    lines.append("")

    for row in preview_rows:
        lines.append("\t".join(row))

    if total_rows > 21:
        lines.append(f"\n... ({total_rows - 21} more rows not shown)")

    return "\n".join(lines)


def write_spreadsheet(input: str) -> str:
    """Create a CSV or Excel file in the sandbox directory.
    Input must be a JSON string with keys:
      - 'filename': e.g. 'output.csv' or 'report.xlsx'
      - 'headers': list of column names
      - 'rows': list of lists
    """
    try:
        data = json.loads(input)
    except json.JSONDecodeError as e:
        return f"Error: input must be valid JSON. {e}"

    filename = data.get("filename", "output.csv")
    headers = data.get("headers", [])
    rows = data.get("rows", [])

    os.makedirs(SANDBOX_DIR, exist_ok=True)
    full_path = os.path.join(SANDBOX_DIR, filename)
    ext = os.path.splitext(filename)[1].lower()

    try:
        if ext == ".csv":
            with open(full_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                if headers:
                    writer.writerow(headers)
                writer.writerows(rows)
        elif ext in (".xlsx", ".xls"):
            wb = openpyxl.Workbook()
            ws = wb.active
            if headers:
                ws.append(headers)
            for row in rows:
                ws.append(row)
            wb.save(full_path)
        else:
            return f"Error: unsupported file type '{ext}'. Use .csv or .xlsx."
    except Exception as e:
        return f"Error writing file: {e}"

    return f"Spreadsheet successfully created at {SANDBOX_DIR}/{filename} ({len(rows)} data rows)"


def chart_data(input: str) -> str:
    """Generate a PNG chart from data and save it to the sandbox directory.
    Input must be a JSON string with keys:
      - 'filename': output filename, e.g. 'chart.png'
      - 'chart_type': one of 'bar', 'line', 'pie', 'scatter'
      - 'title': chart title (optional)
      - 'x_label': x-axis label (optional)
      - 'y_label': y-axis label (optional)
      - 'labels': list of labels
      - 'datasets': list of dataset objects with 'label' and 'values'
    """
    try:
        data = json.loads(input)
    except json.JSONDecodeError as e:
        return f"Error: input must be valid JSON. {e}"

    filename = data.get("filename", "chart.png")
    chart_type = data.get("chart_type", "bar")
    title = data.get("title", "")
    x_label = data.get("x_label", "")
    y_label = data.get("y_label", "")
    labels = data.get("labels", [])
    datasets = data.get("datasets", [])

    if not datasets:
        return "Error: 'datasets' must contain at least one dataset with 'values'."

    os.makedirs(SANDBOX_DIR, exist_ok=True)
    full_path = os.path.join(SANDBOX_DIR, filename)

    try:
        fig, ax = plt.subplots(figsize=(10, 6))

        if chart_type == "pie":
            values = datasets[0].get("values", [])
            ax.pie(values, labels=labels, autopct="%1.1f%%", startangle=90)
            ax.axis("equal")
        elif chart_type == "scatter":
            for ds in datasets:
                values = ds.get("values", [])
                x = list(range(len(values)))
                ax.scatter(x, values, label=ds.get("label", ""))
            if labels:
                ax.set_xticks(range(len(labels)))
                ax.set_xticklabels(labels, rotation=45, ha="right")
        elif chart_type == "line":
            for ds in datasets:
                values = ds.get("values", [])
                x = list(range(len(values)))
                ax.plot(x, values, marker="o", label=ds.get("label", ""))
            if labels:
                ax.set_xticks(range(len(labels)))
                ax.set_xticklabels(labels, rotation=45, ha="right")
        else:  # bar
            import numpy as np
            x = np.arange(len(labels)) if labels else np.arange(len(datasets[0].get("values", [])))
            width = 0.8 / len(datasets)
            for i, ds in enumerate(datasets):
                values = ds.get("values", [])
                offset = (i - len(datasets) / 2 + 0.5) * width
                ax.bar(x + offset, values, width, label=ds.get("label", ""))
            if labels:
                ax.set_xticks(x)
                ax.set_xticklabels(labels, rotation=45, ha="right")

        if title:
            ax.set_title(title)
        if x_label:
            ax.set_xlabel(x_label)
        if y_label:
            ax.set_ylabel(y_label)
        if any(ds.get("label") for ds in datasets) and chart_type != "pie":
            ax.legend()

        plt.tight_layout()
        fig.savefig(full_path, dpi=150)
        plt.close(fig)
    except Exception as e:
        return f"Error creating chart: {e}"

    return f"Chart successfully saved at {SANDBOX_DIR}/{filename}"


def get_tools():
    """Return all document and file tools."""
    file_tools = get_file_tools()

    pdf_tool = Tool(
        name="read_pdf",
        func=read_pdf,
        description="Read and extract text from a PDF file in the sandbox directory. Pass the file path relative to the sandbox folder.",
    )

    create_pdf_tool = Tool(
        name="create_pdf",
        func=create_pdf,
        description=(
            "Create a valid, openable PDF file in the sandbox directory. "
            "ALWAYS use this tool instead of write_file when the output filename ends in .pdf. "
            "Input must be a JSON string with keys: 'filename' (e.g. 'report.pdf'), "
            "'title' (optional heading), and 'content' (the body text). "
            'Example: {"filename": "summary.pdf", "title": "My Title", "content": "Line one\\nLine two"}'
        ),
    )

    read_spreadsheet_tool = Tool(
        name="read_spreadsheet",
        func=read_spreadsheet,
        description=(
            "Read a CSV or Excel (.xlsx) file from the sandbox directory and return a summary with column names and the first rows. "
            "Pass the file path relative to the sandbox folder (e.g. 'data.csv' or 'report.xlsx')."
        ),
    )

    write_spreadsheet_tool = Tool(
        name="write_spreadsheet",
        func=write_spreadsheet,
        description=(
            "Create a CSV or Excel (.xlsx) file in the sandbox directory. "
            "Input must be a JSON string with keys: 'filename' (e.g. 'output.csv'), "
            "'headers' (list of column names), and 'rows' (list of lists). "
            'Example: {"filename": "people.csv", "headers": ["Name", "Age"], "rows": [["Alice", "30"], ["Bob", "25"]]}'
        ),
    )

    chart_data_tool = Tool(
        name="chart_data",
        func=chart_data,
        description=(
            "Generate a PNG chart (bar, line, pie, or scatter) from data and save it to the sandbox directory. "
            "Input must be a JSON string with keys: 'filename', 'chart_type' (bar/line/pie/scatter), "
            "'title' (optional), 'x_label'/'y_label' (optional), 'labels' (list), "
            "and 'datasets' (list of objects with 'label' and 'values')."
        ),
    )

    return file_tools + [pdf_tool, create_pdf_tool, read_spreadsheet_tool, write_spreadsheet_tool, chart_data_tool]
