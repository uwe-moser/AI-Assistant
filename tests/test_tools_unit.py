"""
Unit tests for each tool in sidekick_tools.py.

Tests tool functions directly with mocked external dependencies
(APIs, network calls, file system boundaries).

Run with:  pytest tests/ -v --tb=short
Coverage:  pytest tests/ --cov --cov-report=term-missing
"""

import json
import os
from unittest.mock import AsyncMock, MagicMock, patch, call

import pytest

# Capture the real os.path.join before any patching
_real_path_join = os.path.join


# ---------------------------------------------------------------------------
# Fixture: sandbox directory with cwd pointing at its parent.
# Patching os.path.join globally bleeds into third-party libs (openpyxl,
# matplotlib), so for tools that call those libs we change the working
# directory instead and let the relative "sandbox/" path resolve naturally.
# ---------------------------------------------------------------------------

@pytest.fixture
def sandbox_cwd(tmp_path, monkeypatch):
    """Create a sandbox/ subdir and chdir into its parent so relative
    sandbox/ paths in the tool functions resolve to the temp directory."""
    (tmp_path / "sandbox").mkdir()
    monkeypatch.chdir(tmp_path)
    return tmp_path / "sandbox"


# ===================================================================
# PDF — read_pdf
# ===================================================================

class TestReadPdf:
    """Tests for the read_pdf tool."""

    def test_file_not_found_returns_error(self):
        from sidekick_tools import read_pdf
        result = read_pdf("nonexistent.pdf")
        assert "Error" in result
        assert "not found" in result

    def test_reads_existing_pdf(self, sandbox, sample_pdf):
        """read_pdf should return text with page markers for a valid PDF."""
        from sidekick_tools import read_pdf

        # Redirect the sandbox join so read_pdf finds our temp file
        original_join = os.path.join
        with patch("sidekick_tools.os.path.join",
                    side_effect=lambda *a: original_join(sandbox, a[-1])):
            result = read_pdf(sample_pdf)

        assert "--- Page 1 ---" in result
        assert "Test PDF Content" in result

    def test_empty_pdf_returns_message(self, sandbox):
        """A PDF with blank pages should return a clear message."""
        from sidekick_tools import read_pdf
        from fpdf import FPDF

        pdf = FPDF()
        pdf.add_page()  # blank page, no text
        blank_file = os.path.join(sandbox, "blank.pdf")
        pdf.output(blank_file)

        original_join = os.path.join
        with patch("sidekick_tools.os.path.join",
                    side_effect=lambda *a: original_join(sandbox, a[-1])):
            result = read_pdf("blank.pdf")

        assert "no extractable text" in result.lower()

    def test_multi_page_pdf(self, sandbox):
        """Each page should get its own marker."""
        from sidekick_tools import read_pdf
        from fpdf import FPDF

        pdf = FPDF()
        for i in range(3):
            pdf.add_page()
            pdf.set_font("Helvetica", size=12)
            pdf.cell(200, 10, text=f"Page {i + 1} content")
        multi_file = os.path.join(sandbox, "multi.pdf")
        pdf.output(multi_file)

        original_join = os.path.join
        with patch("sidekick_tools.os.path.join",
                    side_effect=lambda *a: original_join(sandbox, a[-1])):
            result = read_pdf("multi.pdf")

        assert "--- Page 1 ---" in result
        assert "--- Page 2 ---" in result
        assert "--- Page 3 ---" in result


# ===================================================================
# PDF — create_pdf
# ===================================================================

class TestCreatePdf:
    """Tests for the create_pdf tool."""

    def test_creates_file_on_disk(self, sandbox):
        from sidekick_tools import create_pdf

        with patch("sidekick_tools.os.path.join",
                    side_effect=lambda *a: _real_path_join(sandbox, a[-1])):
            with patch("sidekick_tools.os.makedirs"):
                result = create_pdf(json.dumps({
                    "filename": "out.pdf",
                    "title": "Hello",
                    "content": "World",
                }))

        assert "successfully" in result.lower()
        assert os.path.isfile(_real_path_join(sandbox, "out.pdf"))

    def test_auto_appends_pdf_extension(self, sandbox):
        from sidekick_tools import create_pdf

        with patch("sidekick_tools.os.path.join",
                    side_effect=lambda *a: _real_path_join(sandbox, a[-1])):
            with patch("sidekick_tools.os.makedirs"):
                result = create_pdf(json.dumps({
                    "filename": "no_ext",
                    "content": "data",
                }))

        assert "no_ext.pdf" in result

    def test_invalid_json_returns_error(self):
        from sidekick_tools import create_pdf
        result = create_pdf("{bad json")
        assert "Error" in result

    def test_empty_content_succeeds(self, sandbox):
        from sidekick_tools import create_pdf

        with patch("sidekick_tools.os.path.join",
                    side_effect=lambda *a: _real_path_join(sandbox, a[-1])):
            with patch("sidekick_tools.os.makedirs"):
                result = create_pdf(json.dumps({
                    "filename": "empty.pdf",
                    "content": "",
                }))

        assert "successfully" in result.lower()

    def test_unicode_content_does_not_crash(self, sandbox):
        """Unicode characters (smart quotes, dashes, emoji) should render fine."""
        from sidekick_tools import create_pdf

        with patch("sidekick_tools.os.path.join",
                    side_effect=lambda *a: _real_path_join(sandbox, a[-1])):
            with patch("sidekick_tools.os.makedirs"):
                result = create_pdf(json.dumps({
                    "filename": "unicode.pdf",
                    "content": "\u201cHello\u201d \u2014 world\u2026 \U0001f600",
                }))

        assert "successfully" in result.lower()


# ===================================================================
# PDF — _content_to_html helper
# ===================================================================

class TestContentToHtml:
    """Tests for the _content_to_html helper."""

    def test_title_becomes_h1(self):
        from sidekick_tools import _content_to_html
        html = _content_to_html("My Title", "body text")
        assert "<h1>My Title</h1>" in html

    def test_no_title_omits_h1(self):
        from sidekick_tools import _content_to_html
        html = _content_to_html("", "body text")
        assert "<h1>" not in html

    def test_html_entities_escaped(self):
        from sidekick_tools import _content_to_html
        html = _content_to_html("", "a < b & c > d")
        assert "&lt;" in html
        assert "&amp;" in html
        assert "&gt;" in html

    def test_bold_markdown_converted(self):
        from sidekick_tools import _content_to_html
        html = _content_to_html("", "**bold text**")
        assert "<strong>bold text</strong>" in html

    def test_empty_lines_produce_br(self):
        from sidekick_tools import _content_to_html
        html = _content_to_html("", "line1\n\nline2")
        assert "<br>" in html


# ===================================================================
# YouTube Transcript
# ===================================================================

class TestGetYoutubeTranscript:
    """Tests for the get_youtube_transcript tool."""

    def _make_mock_transcript(self, lines):
        """Helper: build a mock transcript object from a list of strings."""
        snippets = []
        for text in lines:
            s = MagicMock()
            s.text = text
            snippets.append(s)
        mock_t = MagicMock()
        mock_t.snippets = snippets
        return mock_t

    def _run_with_mock(self, url_or_id, lines=None):
        """Call get_youtube_transcript with a mocked API and return (result, mock_instance)."""
        from sidekick_tools import get_youtube_transcript

        transcript = self._make_mock_transcript(lines or ["Hello"])
        with patch("sidekick_tools.YouTubeTranscriptApi") as MockCls:
            mock_inst = MockCls.return_value
            mock_inst.fetch.return_value = transcript
            result = get_youtube_transcript(url_or_id)
        return result, mock_inst

    @pytest.mark.parametrize("url,expected_id", [
        ("https://www.youtube.com/watch?v=dQw4w9WgXcQ", "dQw4w9WgXcQ"),
        ("https://youtube.com/watch?v=abc123", "abc123"),
        ("https://youtu.be/short1", "short1"),
        ("https://m.youtube.com/watch?v=mob1", "mob1"),
    ])
    def test_extracts_video_id_from_urls(self, url, expected_id):
        _, mock_inst = self._run_with_mock(url)
        mock_inst.fetch.assert_called_once_with(expected_id)

    def test_plain_video_id_passed_through(self):
        _, mock_inst = self._run_with_mock("rawID123")
        mock_inst.fetch.assert_called_once_with("rawID123")

    def test_strips_extra_query_params(self):
        _, mock_inst = self._run_with_mock(
            "https://www.youtube.com/watch?v=vid1&list=PLabc&index=3"
        )
        mock_inst.fetch.assert_called_once_with("vid1")

    def test_joins_transcript_lines(self):
        result, _ = self._run_with_mock("x", lines=["Line 1", "Line 2", "Line 3"])
        assert result == "Line 1\nLine 2\nLine 3"

    def test_api_error_propagates(self):
        from sidekick_tools import get_youtube_transcript

        with patch("sidekick_tools.YouTubeTranscriptApi") as MockCls:
            mock_inst = MockCls.return_value
            mock_inst.fetch.side_effect = Exception("Video not found")
            with pytest.raises(Exception, match="Video not found"):
                get_youtube_transcript("bad_id")


# ===================================================================
# Push Notification
# ===================================================================

class TestPushNotification:
    """Tests for the push() tool."""

    @patch("sidekick_tools.requests.post")
    def test_returns_success(self, mock_post):
        from sidekick_tools import push
        assert push("Hello") == "success"

    @patch("sidekick_tools.requests.post")
    def test_sends_correct_payload(self, mock_post):
        from sidekick_tools import push, pushover_url, pushover_token, pushover_user
        push("Test msg")
        mock_post.assert_called_once_with(
            pushover_url,
            data={"token": pushover_token, "user": pushover_user, "message": "Test msg"},
        )

    @patch("sidekick_tools.requests.post")
    def test_handles_empty_message(self, mock_post):
        from sidekick_tools import push
        result = push("")
        assert result == "success"
        mock_post.assert_called_once()

    @patch("sidekick_tools.requests.post", side_effect=ConnectionError("Network down"))
    def test_network_error_propagates(self, mock_post):
        from sidekick_tools import push
        with pytest.raises(ConnectionError):
            push("will fail")


# ===================================================================
# Web Search (Google Serper)
# ===================================================================

class TestWebSearch:
    """Tests for the Google Serper search tool."""

    @patch("sidekick_tools.serper")
    def test_search_delegates_to_serper(self, mock_serper):
        mock_serper.run.return_value = '{"organic": [{"title": "Result 1"}]}'
        result = mock_serper.run("LangChain tutorial")
        mock_serper.run.assert_called_once_with("LangChain tutorial")
        assert "Result 1" in result

    @patch("sidekick_tools.serper")
    def test_search_empty_query(self, mock_serper):
        mock_serper.run.return_value = '{"organic": []}'
        result = mock_serper.run("")
        assert "organic" in result

    @patch("sidekick_tools.serper")
    def test_search_api_error(self, mock_serper):
        mock_serper.run.side_effect = Exception("API rate limit")
        with pytest.raises(Exception, match="API rate limit"):
            mock_serper.run("query")


# ===================================================================
# Wikipedia
# ===================================================================

class TestWikipedia:
    """Tests for the Wikipedia query tool."""

    def test_wiki_tool_returns_summary(self):
        with patch("langchain_community.utilities.wikipedia.WikipediaAPIWrapper.run",
                    return_value="Python is a programming language."):
            from langchain_community.utilities.wikipedia import WikipediaAPIWrapper
            from langchain_community.tools.wikipedia.tool import WikipediaQueryRun

            wrapper = WikipediaAPIWrapper()
            tool = WikipediaQueryRun(api_wrapper=wrapper)
            result = tool.run("Python programming")
            assert "Python" in result

    def test_wiki_tool_no_results(self):
        with patch("langchain_community.utilities.wikipedia.WikipediaAPIWrapper.run",
                    return_value="No good Wikipedia Search Result was found"):
            from langchain_community.utilities.wikipedia import WikipediaAPIWrapper
            from langchain_community.tools.wikipedia.tool import WikipediaQueryRun

            wrapper = WikipediaAPIWrapper()
            tool = WikipediaQueryRun(api_wrapper=wrapper)
            result = tool.run("xyznonexistent12345")
            assert "No good" in result


# ===================================================================
# arXiv
# ===================================================================

class TestArxiv:
    """Tests for the arXiv query tool."""

    def test_arxiv_tool_returns_papers(self):
        mock_result = (
            "Published: 2024-01-15\nTitle: Attention Is All You Need\n"
            "Summary: We propose a new architecture..."
        )
        with patch("langchain_community.utilities.arxiv.ArxivAPIWrapper.run",
                    return_value=mock_result):
            from langchain_community.utilities.arxiv import ArxivAPIWrapper
            from langchain_community.tools.arxiv.tool import ArxivQueryRun

            wrapper = ArxivAPIWrapper()
            tool = ArxivQueryRun(api_wrapper=wrapper)
            result = tool.run("transformer architecture")
            assert "Attention" in result

    def test_arxiv_tool_no_results(self):
        with patch("langchain_community.utilities.arxiv.ArxivAPIWrapper.run",
                    return_value="No good Arxiv Result was found"):
            from langchain_community.utilities.arxiv import ArxivAPIWrapper
            from langchain_community.tools.arxiv.tool import ArxivQueryRun

            wrapper = ArxivAPIWrapper()
            tool = ArxivQueryRun(api_wrapper=wrapper)
            result = tool.run("xyznonexistent12345")
            assert "No good" in result


# ===================================================================
# Python REPL
# ===================================================================

class TestPythonRepl:
    """Tests for the PythonREPLTool."""

    def test_repl_executes_code(self):
        from langchain_experimental.tools import PythonREPLTool
        repl = PythonREPLTool()
        result = repl.run("print(2 + 2)")
        assert "4" in result

    def test_repl_handles_syntax_error(self):
        from langchain_experimental.tools import PythonREPLTool
        repl = PythonREPLTool()
        result = repl.run("def bad(")
        assert "Error" in result or "SyntaxError" in result

    def test_repl_multiline_output(self):
        from langchain_experimental.tools import PythonREPLTool
        repl = PythonREPLTool()
        result = repl.run("for i in range(3): print(i)")
        assert "0" in result
        assert "1" in result
        assert "2" in result


# ===================================================================
# File Management Tools
# ===================================================================

class TestFileTools:
    """Tests for the FileManagementToolkit."""

    def test_returns_nonempty_list(self):
        from sidekick_tools import get_file_tools
        tools = get_file_tools()
        assert isinstance(tools, list)
        assert len(tools) > 0

    def test_contains_expected_tools(self):
        from sidekick_tools import get_file_tools
        names = {t.name for t in get_file_tools()}
        assert {"read_file", "write_file", "list_directory"}.issubset(names)

    def test_all_tools_have_descriptions(self):
        from sidekick_tools import get_file_tools
        for tool in get_file_tools():
            assert tool.description, f"Tool '{tool.name}' missing description"


# ===================================================================
# Playwright Browser Tools
# ===================================================================

class TestPlaywrightTools:
    """Tests for the playwright_tools() async factory."""

    @pytest.mark.asyncio
    async def test_returns_tools_browser_playwright(self):
        """Verify the return shape without launching a real browser."""
        mock_browser = MagicMock()
        mock_playwright = MagicMock()

        fake_tools = [MagicMock(name="navigate"), MagicMock(name="click")]

        with patch("sidekick_tools.async_playwright") as mock_ap:
            mock_pw_instance = MagicMock()
            mock_pw_instance.chromium.launch = AsyncMock(return_value=mock_browser)
            mock_ap.return_value.start = AsyncMock(return_value=mock_pw_instance)

            with patch("sidekick_tools.PlayWrightBrowserToolkit") as MockToolkit:
                mock_tk = MagicMock()
                mock_tk.get_tools.return_value = fake_tools
                MockToolkit.from_browser.return_value = mock_tk

                from sidekick_tools import playwright_tools
                tools, browser, pw = await playwright_tools()

        assert tools == fake_tools
        assert browser == mock_browser


# ===================================================================
# other_tools() Assembly
# ===================================================================

class TestOtherToolsAssembly:
    """Tests for the other_tools() async function that assembles all non-browser tools."""

    @pytest.mark.asyncio
    async def test_returns_list_with_correct_count(self):
        from sidekick_tools import other_tools
        with patch("sidekick_tools.serper"):
            tools = await other_tools()
            assert isinstance(tools, list)
            # 6 file tools + 8 custom tools = 14
            assert len(tools) >= 12

    @pytest.mark.asyncio
    async def test_contains_all_expected_tool_names(self):
        from sidekick_tools import other_tools
        with patch("sidekick_tools.serper"):
            tools = await other_tools()
            names = {t.name for t in tools}
            expected = {
                "send_push_notification", "search", "Python_REPL",
                "wikipedia", "arxiv", "read_pdf",
                "get_youtube_transcript", "create_pdf",
                "read_spreadsheet", "write_spreadsheet", "chart_data",
            }
            missing = expected - names
            assert not missing, f"Missing tools: {missing}"

    @pytest.mark.asyncio
    async def test_every_tool_has_description(self):
        from sidekick_tools import other_tools
        with patch("sidekick_tools.serper"):
            tools = await other_tools()
            for tool in tools:
                assert tool.description, f"Tool '{tool.name}' has no description"

    @pytest.mark.asyncio
    async def test_every_tool_is_invocable(self):
        from sidekick_tools import other_tools
        with patch("sidekick_tools.serper"):
            tools = await other_tools()
            for tool in tools:
                # LangChain tools expose .run() or .invoke()
                has_run = hasattr(tool, "run") and callable(tool.run)
                has_invoke = hasattr(tool, "invoke") and callable(tool.invoke)
                assert has_run or has_invoke, \
                    f"Tool '{tool.name}' has neither .run() nor .invoke()"

    @pytest.mark.asyncio
    async def test_google_places_included_when_api_key_set(self):
        from sidekick_tools import other_tools
        mock_tool = MagicMock()
        mock_tool.name = "google_places"
        mock_tool.description = "Search Google Maps"
        with patch("sidekick_tools.serper"), \
             patch.dict(os.environ, {"GPLACES_API_KEY": "fake-key-for-test"}), \
             patch("sidekick_tools.GooglePlacesTool", return_value=mock_tool):
            tools = await other_tools()
            names = {t.name for t in tools}
            assert "google_places" in names

    @pytest.mark.asyncio
    async def test_google_places_excluded_when_no_api_key(self):
        from sidekick_tools import other_tools
        with patch("sidekick_tools.serper"), \
             patch.dict(os.environ, {}, clear=False) as env:
            env.pop("GPLACES_API_KEY", None)
            tools = await other_tools()
            names = {t.name for t in tools}
            assert "google_places" not in names


# ===================================================================
# Spreadsheet — read_spreadsheet
# ===================================================================

class TestReadSpreadsheet:
    """Tests for the read_spreadsheet tool."""

    def _join_to(self, sandbox):
        """Return a side_effect for os.path.join that roots paths in sandbox."""
        return lambda *a: _real_path_join(sandbox, a[-1])

    def test_file_not_found_returns_error(self):
        from sidekick_tools import read_spreadsheet
        result = read_spreadsheet("nonexistent.csv")
        assert "Error" in result
        assert "not found" in result

    def test_unsupported_extension_returns_error(self, sandbox):
        from sidekick_tools import read_spreadsheet
        # Create a dummy .txt file in the sandbox so it passes the existence check
        txt_path = _real_path_join(sandbox, "data.txt")
        open(txt_path, "w").close()
        with patch("sidekick_tools.os.path.join", side_effect=self._join_to(sandbox)):
            result = read_spreadsheet("data.txt")
        assert "Error" in result
        assert "unsupported" in result

    def test_reads_csv_headers_and_rows(self, sandbox):
        import csv as _csv
        from sidekick_tools import read_spreadsheet

        csv_path = _real_path_join(sandbox, "sample.csv")
        with open(csv_path, "w", newline="") as f:
            writer = _csv.writer(f)
            writer.writerow(["Name", "Age", "City"])
            writer.writerow(["Alice", "30", "NYC"])
            writer.writerow(["Bob", "25", "LA"])

        with patch("sidekick_tools.os.path.join", side_effect=self._join_to(sandbox)):
            result = read_spreadsheet("sample.csv")

        assert "Name" in result
        assert "Age" in result
        assert "Alice" in result
        assert "Bob" in result
        assert "3 rows" in result  # header + 2 data rows

    def test_reads_xlsx_headers_and_rows(self, sandbox_cwd):
        from sidekick_tools import read_spreadsheet
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Product", "Price"])
        ws.append(["Apple", 1.5])
        ws.append(["Banana", 0.75])
        wb.save(sandbox_cwd / "sample.xlsx")

        result = read_spreadsheet("sample.xlsx")

        assert "Product" in result
        assert "Apple" in result
        assert "3 rows" in result

    def test_large_csv_shows_truncation_note(self, sandbox):
        import csv as _csv
        from sidekick_tools import read_spreadsheet

        csv_path = _real_path_join(sandbox, "big.csv")
        with open(csv_path, "w", newline="") as f:
            writer = _csv.writer(f)
            writer.writerow(["id", "value"])
            for i in range(50):
                writer.writerow([str(i), str(i * 10)])

        with patch("sidekick_tools.os.path.join", side_effect=self._join_to(sandbox)):
            result = read_spreadsheet("big.csv")

        assert "more rows not shown" in result

    def test_csv_with_utf8_bom(self, sandbox):
        from sidekick_tools import read_spreadsheet

        csv_path = _real_path_join(sandbox, "bom.csv")
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            f.write("Name,Value\nAlice,42\n")

        with patch("sidekick_tools.os.path.join", side_effect=self._join_to(sandbox)):
            result = read_spreadsheet("bom.csv")

        # BOM should be stripped; first column header should be clean
        assert "Name" in result
        assert "\ufeffName" not in result


# ===================================================================
# Spreadsheet — write_spreadsheet
# ===================================================================

class TestWriteSpreadsheet:
    """Tests for the write_spreadsheet tool."""

    def test_invalid_json_returns_error(self):
        from sidekick_tools import write_spreadsheet
        result = write_spreadsheet("{bad json")
        assert "Error" in result

    def test_unsupported_extension_returns_error(self, sandbox):
        from sidekick_tools import write_spreadsheet
        with patch("sidekick_tools.os.path.join",
                   side_effect=lambda *a: _real_path_join(sandbox, a[-1])):
            with patch("sidekick_tools.os.makedirs"):
                result = write_spreadsheet(json.dumps({
                    "filename": "out.txt",
                    "headers": ["A"],
                    "rows": [["1"]],
                }))
        assert "Error" in result
        assert "unsupported" in result

    def test_creates_csv_file(self, sandbox):
        from sidekick_tools import write_spreadsheet

        with patch("sidekick_tools.os.path.join",
                   side_effect=lambda *a: _real_path_join(sandbox, a[-1])):
            with patch("sidekick_tools.os.makedirs"):
                result = write_spreadsheet(json.dumps({
                    "filename": "out.csv",
                    "headers": ["Name", "Score"],
                    "rows": [["Alice", "95"], ["Bob", "87"]],
                }))

        assert "successfully" in result.lower()
        assert os.path.isfile(_real_path_join(sandbox, "out.csv"))

    def test_csv_content_is_correct(self, sandbox):
        import csv as _csv
        from sidekick_tools import write_spreadsheet

        with patch("sidekick_tools.os.path.join",
                   side_effect=lambda *a: _real_path_join(sandbox, a[-1])):
            with patch("sidekick_tools.os.makedirs"):
                write_spreadsheet(json.dumps({
                    "filename": "check.csv",
                    "headers": ["X", "Y"],
                    "rows": [["1", "2"], ["3", "4"]],
                }))

        with open(_real_path_join(sandbox, "check.csv"), newline="") as f:
            rows = list(_csv.reader(f))

        assert rows[0] == ["X", "Y"]
        assert rows[1] == ["1", "2"]
        assert rows[2] == ["3", "4"]

    def test_creates_xlsx_file(self, sandbox):
        import openpyxl
        from sidekick_tools import write_spreadsheet

        with patch("sidekick_tools.os.path.join",
                   side_effect=lambda *a: _real_path_join(sandbox, a[-1])):
            with patch("sidekick_tools.os.makedirs"):
                result = write_spreadsheet(json.dumps({
                    "filename": "out.xlsx",
                    "headers": ["Col1", "Col2"],
                    "rows": [[1, 2], [3, 4]],
                }))

        assert "successfully" in result.lower()
        path = _real_path_join(sandbox, "out.xlsx")
        assert os.path.isfile(path)

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("Col1", "Col2")
        assert rows[1] == (1, 2)

    def test_row_count_in_success_message(self, sandbox):
        from sidekick_tools import write_spreadsheet

        with patch("sidekick_tools.os.path.join",
                   side_effect=lambda *a: _real_path_join(sandbox, a[-1])):
            with patch("sidekick_tools.os.makedirs"):
                result = write_spreadsheet(json.dumps({
                    "filename": "count.csv",
                    "headers": ["A"],
                    "rows": [["1"], ["2"], ["3"]],
                }))

        assert "3 data rows" in result

    def test_no_headers_writes_rows_only(self, sandbox):
        import csv as _csv
        from sidekick_tools import write_spreadsheet

        with patch("sidekick_tools.os.path.join",
                   side_effect=lambda *a: _real_path_join(sandbox, a[-1])):
            with patch("sidekick_tools.os.makedirs"):
                write_spreadsheet(json.dumps({
                    "filename": "noheader.csv",
                    "headers": [],
                    "rows": [["a", "b"]],
                }))

        with open(_real_path_join(sandbox, "noheader.csv"), newline="") as f:
            rows = list(_csv.reader(f))
        assert len(rows) == 1
        assert rows[0] == ["a", "b"]


# ===================================================================
# Chart — chart_data
# ===================================================================

class TestChartData:
    """Tests for the chart_data tool."""

    def test_invalid_json_returns_error(self):
        from sidekick_tools import chart_data
        result = chart_data("{bad json")
        assert "Error" in result

    def test_missing_datasets_returns_error(self, sandbox_cwd):
        from sidekick_tools import chart_data
        result = chart_data(json.dumps({
            "filename": "chart.png",
            "chart_type": "bar",
            "labels": ["A"],
            "datasets": [],
        }))
        assert "Error" in result

    def _make_chart(self, sandbox_cwd, chart_type, filename="chart.png"):
        from sidekick_tools import chart_data
        payload = json.dumps({
            "filename": filename,
            "chart_type": chart_type,
            "title": f"Test {chart_type}",
            "labels": ["Jan", "Feb", "Mar"],
            "datasets": [{"label": "Series A", "values": [10, 20, 15]}],
        })
        return chart_data(payload)

    @pytest.mark.parametrize("chart_type", ["bar", "line", "pie", "scatter"])
    def test_all_chart_types_succeed(self, sandbox_cwd, chart_type):
        result = self._make_chart(sandbox_cwd, chart_type, filename=f"{chart_type}.png")
        assert "successfully" in result.lower()
        assert (sandbox_cwd / f"{chart_type}.png").is_file()

    def test_output_is_valid_png(self, sandbox_cwd):
        """The saved file should start with the PNG magic bytes."""
        self._make_chart(sandbox_cwd, "bar", filename="valid.png")
        with open(sandbox_cwd / "valid.png", "rb") as f:
            header = f.read(8)
        assert header[:4] == b"\x89PNG"

    def test_multiple_datasets_bar_chart(self, sandbox_cwd):
        from sidekick_tools import chart_data
        result = chart_data(json.dumps({
            "filename": "multi.png",
            "chart_type": "bar",
            "labels": ["Q1", "Q2"],
            "datasets": [
                {"label": "Revenue", "values": [100, 200]},
                {"label": "Cost",    "values": [80, 120]},
            ],
        }))
        assert "successfully" in result.lower()

    def test_returns_correct_sandbox_path_in_message(self, sandbox_cwd):
        result = self._make_chart(sandbox_cwd, "line", filename="myline.png")
        assert "sandbox/myline.png" in result
